# -*- coding: utf-8 -*-
__author__ = 'Administrator'
"""
@file: utilsDemo.py
@time: 2017/12/13 15:18
"""
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter




def read():
    wordbook = load_workbook('xxx.xlsx')
    word_sheet = wordbook['title']

    row_cell = tuple(word_sheet.rows)
    row_max_row = len(row_cell)
    row_max_col = len(row_cell[0])

    content = word_sheet.iter_rows(min_row=1, max_row=row_max_row,
                                   max_col=row_max_col)

    import pandas as pd
    from openpyxl.utils.dataframe import dataframe_to_rows

    wk = Workbook()
    ws = wk.active
    ws.title = 'ran'
    import datetime
    now = datetime.datetime.now()
    dates = pd.date_range(now, periods=row_max_row)
    # 转换
    df = pd.DataFrame(content, index=dates, columns=list(range(row_max_col)))
    row_true = True
    for row in dataframe_to_rows(df, index=True, header=True):
        list1 = []
        if row_true:
            pass
        else:
            print(row)
            for r in range(1, len(row)):
                list1.append(row[r].value)
            ws.append(list1)
        row_true = False

    ws['F6'] = now
    print(ws['F6'].number_format)
    wk.save('yyy.xlsx')




if __name__ == '__main__':
    read()



