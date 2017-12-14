# -*- coding: utf-8 -*- 
"""
@__author__ :70486 
@file: OpenpyxlExcel.py
@time: 2017/12/9 18:39
@项目名称:operating
http://openpyxl.readthedocs.io/en/latest/
http://openpyxl.readthedocs.io/en/default/usage.html#write-a-workbook
http://blog.csdn.net/tanzuozhev/article/details/76713387
http://www.cnblogs.com/chaosimple/p/4153083.html
"""

from openpyxl import Workbook,load_workbook
import datetime
import os

"""
这个class负责读取数据：从excel中读取数据到电脑
"""


class READEXCEL:
    def __init__(self, FILEPATH, SHEETNAME=0):
        """
            FILEPATH :需要文件的位置
            SHEETNAME　：读取工作薄的页面或者工作薄名
        """
        # 判断文件是否存在
        if os.path.exists(FILEPATH):
            self.excel = FILEPATH

            # 判断是否为空
            self.data(SHEETNAME)
        else:
            raise FileNotFoundError('文件不存在！')

    def data(self, sheet):#　此处不严谨：如果输入的字符串都为数字那么就会出错

        # 创建需要操作的文档
        self.workbook = load_workbook(filename=self.excel)  # 打开文档

        # 判断是根据数字还是文字进行读取sheet，如果是数字的话必须小于现有的长度
        if type(sheet) in [int] and sheet < len(self.workbook.sheetnames):
            self.sheetbook = self.workbook[self.workbook.sheetnames[sheet]]

        # 如果是文字
        elif type(sheet) in [str]:
            self.sheetbook = self.workbook[self.workbook.sheetnames[sheet]]

        # 长度过长时提示
        elif sheet >= len(self.workbook.sheetnames):
            print("sheet索要的位置大于现有的长度")

        # 最后输出
        else:
            print("你输入啥咯.")

    #  返回读取到单元格的内容，并已cell形式返回
    def get_sheet_value(self, _value):
        # 先判断需要寻找的cell位置。如果为空或者其他类型的就提示
        if type(_value) in [int, str]:
            content = self.sheetbook[_value]
            # 打印指定的内容:ws['A4']返回的是一个cell，通过value来获取值
            # content = self.sheetbook[_value].value
            return content
        else:
            print('1级错误')

    #   指定行和列来读取数据信息
    def position_sheet_value(self,min_row=1, max_row=1, max_col=1 ):
        content = self.sheetbook.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col)
        return content

    #   工作薄的名称
    def get_sheet_title(self):
        return self.sheetbook.title

    #   返回某个列的标题位置名称
    def get_column_letter(self,Number = 1):
        from openpyxl.utils import get_column_letter
        return get_column_letter(Number)

    #   返回当前工作薄的复制体对象
    def replica_worksheet(self):
        copy_sheet = self.workbook.copy_worksheet(self.sheetbook)
        return copy_sheet

    #   为真时，以行为一体，每行的数据信息
    #   为假时，以列为一体，每列的数据信息
    def total_row_columns(self,total = True):
        if total:
            content = tuple(read_excel.sheetbook.rows)# 单行中，列的长度
        else:
            content = tuple(read_excel.sheetbook.columns)# 单列中，行的长度
        return content;

    """
    将现有xlsx文档保存为xltx模板
    """
    def attribute_template(self,emplate = None):
        # 将现有的表进行复制并保存为模板。。并后缀名为xltx，如果为xls和xlsx在打开的时候出现问题
        if type(emplate) in [str] :#  判断输入的内容是否为字符串

            genericpath = os.path.splitext(emplate)[1] #    切割文件后缀名

            if genericpath =='.xltx':

                attribute = os.path.split(emplate)[1] # 切割最后一个文件的名字
                self.workbook.template = True   # 属性设置
                self.workbook.save(attribute)   # 保存后缀名为xltx的文件
                print('The xlsx document is completed by turning the xitx template.')
                return attribute;   # 返回文件名

            elif genericpath == '':
                attribute = os.path.split(emplate)[1] + '.xltx'# 切割最后一个文件的名字
                self.workbook.template = True# 属性设置
                self.workbook.save(attribute)# 保存后缀名为xltx的文件
                print('The xlsx document is completed by turning the xitx template.')
                return attribute; # 返回文件名

            elif genericpath in ['xls','xlsx','txt']:
                print('Files that do not support suffixes such as XLS xlsx text')

            else:
                print('The input file suffix name does not conform.')

        elif emplate == None:
            path = os.path.splitext(os.path.split(self.excel)[1])[0]
            attribute = 'copy_' + path + '.xltx'  # 设置文件的名字
            self.workbook.template = True
            self.workbook.save(attribute)
            print('Xlsx turns xltx file.')
            return attribute

        else:
            print('你丫的文件输入有误')

    """
    将现有xltx模板转成xlsx文档进行保存
    """
    def attribute_document(self,template,document):
        # 将现有的模板还原成文档或直接将现有的wb另存为。
        # 保存为xls文件打开的时候会提示错误
        if os.path.exists(template): # 判断文件是否存在

            if os.path.splitext(template)[1] == '.xltx' : # 判断模板是不是xltx文件
                attribute = os.path.split(document)  # 将文档切割。切成路径和文件两部分

                if attribute[0] == '': # 判断路径是否为空，为空说明保存跟模板同一个位置

                    if os.path.splitext(attribute[1])[1] == '.xlsx': # 判断是否文档保存是否为xlsx文件
                        self.workbook.template = False
                        self.workbook.save(document)
                        print('The xitx template turns to the xlsx document.')
                        return document  # 返回文件名
                    else:
                        print(document + ' : Not a File with a suffix xlsx')
                elif os.path.exists(attribute[0]): #　有路径说明要保存在指定路径下面

                    if os.path.splitext(attribute[1])[1] == '.xlsx':

                        self.workbook.template = False
                        self.workbook.save(document)
                        print('The xitx template turns to the xlsx document.')
                        return document  # 返回文件名
                    else:
                        print(document + ' : Not a File with a suffix xlsx')
                else:
                    print(document + ' : File path does not exist, please try again.')
            else :
                print(template + ' : Not a File with a suffix xltx')
        else:
            print(template + ' : File does not exist')

    def qisahife(self):
        #　打印工作薄名称
        print(wb.sheetnames)


"""
这个class负责写入数据：将信息从电脑写到excel
"""


class WRITEEXCEL:
    def __init__(self, FILEPATH, SHEETTITLE='title'):
        """
        :param FILEPATH:  需要操作的文件路径
        :param SHEETNAME: 工作薄名称
        """
        # 判断保存文件的位置是否存在
        path = os.path.split(FILEPATH)
        if os.path.exists(path[0]) or path[0] =='' :
            if os.path.splitext(path[1])[1] == '.xlsx':
                self.excel = FILEPATH  # 存储文件

                self.data(SHEETTITLE)
            else:
                print(os.path.splitext(path[1])[1] + ' : 读取文件的格式不对')
        else:
            raise FileNotFoundError('文件不存在！')

    def data(self, sheet):#　此处不严谨：如果输入的字符串都为数字那么就会出错

        # 创建需要操作的文档
        self.workbook = Workbook()
        self.work_sheet  = self.workbook.active # 打开工作薄
        self.work_sheet.title = sheet # 设置工作薄的名字
        self.work_sheet['A1'] = '大爷'
        self.workbook.save(filename=self.excel)
        print('333')



if __name__ == '__main__':
    # read_excel = READEXCEL('empty_book.xlsx')
    # print(read_excel.get_sheet_value('1'))
    #　print(read_excel.attribute_document('copy_empty_book.xlsx.xltx','E://operating//practical//utils//xyy.xlsx'))
    print(datetime.datetime.now())
    write_excel = WRITEEXCEL('E://drivers//xxx.xlsx')
    print(datetime.datetime.now())


